from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.views import LoginView
from django.contrib.admin import AdminSite, ModelAdmin
from django.urls import path
from django.http import HttpResponse
from django.utils.encoding import smart_str
from .models import CustomUser, Stall, MenuItem, Voucher, Order, CartItem, OrderItem
from .forms import SignUpForm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import datetime
import csv

class JSECAdminSite(AdminSite):
    site_header = "JSEC Express Admin"
    site_title = "JSEC Admin"
    index_title = "Welcome to JSEC Express Admin"
    login_template = "core/admin.html"
    index_template = "core/index.html"
    base_template = "core/admin_base_site.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("login/", self.admin_view(LoginView.as_view(template_name=self.login_template)), name="login"),
        ]
        return custom_urls + urls

admin_site = JSECAdminSite(name="jsecadmin")

class CustomUserAdmin(UserAdmin):
    model = CustomUser
    add_form = SignUpForm
    list_display = ("student_id", "full_name", "email", "is_staff")
    fieldsets = (
        (None, {"fields": ("student_id", "password")}),
        ("Personal info", {"fields": ("full_name", "contact_number", "email")}),
        ("Permissions", {"fields": ("is_staff", "is_active","blacklisted")}),
    )
    add_fieldsets = (
        (None, {
            "fields": (
                "student_id", "full_name", "contact_number",
                "email", "password1", "password2"
            )
        }),
    )
    ordering = ("student_id",)

class BaseStallScopedAdmin(ModelAdmin):
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if not request.user.is_authenticated:
            return qs.none()
        if request.user.is_superuser:
            return qs
        if hasattr(self.model, "stall"):
            return qs.filter(stall__owner=request.user)
        if hasattr(self.model, "owner"):
            return qs.filter(owner=request.user)
        return qs.none()

    def has_module_permission(self, request):
        return request.user.is_superuser or self.get_queryset(request).exists()

    def has_view_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        if not request.user.is_authenticated:
            return False
        if obj is None:
            return self.get_queryset(request).exists()
        if hasattr(obj, "stall"):
            return obj.stall.owner == request.user
        if hasattr(obj, "owner"):
            return obj.owner == request.user
        return False

    def has_change_permission(self, request, obj=None):
        return self.has_view_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        return self.has_view_permission(request, obj)

    def has_add_permission(self, request):
        return request.user.is_authenticated

class StallAdmin(BaseStallScopedAdmin):
    list_display = ["name", "owner"]

class MenuItemAdmin(BaseStallScopedAdmin):
    list_display = ["name", "stall", "price"]

class VoucherAdmin(BaseStallScopedAdmin):
    list_display = ["code", "stall", "discount_amount", "is_active"]

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "stall" and request.user.is_authenticated and not request.user.is_superuser:
            kwargs["queryset"] = Stall.objects.filter(owner=request.user)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

class CartItemAdmin(BaseStallScopedAdmin):
    list_display = ["user", "item", "quantity"]

class OrderItemAdmin(BaseStallScopedAdmin):
    list_display = ["order", "item", "quantity"]

class OrderAdmin(BaseStallScopedAdmin):
    list_display = ["user", "stall", "pickup_time", "total_cost", "transaction_id", "is_complete"]
    actions = ["export_orders_as_csv", "export_orders_by_day_excel"]

    def export_orders_as_csv(self, request, queryset):
        if not request.user.is_superuser:
            queryset = queryset.filter(stall__owner=request.user)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=orders_export.csv'

        writer = csv.writer(response)
        writer.writerow([
            "Order ID", "User", "Stall", "Pickup Time", 
            "Total Cost", "Transaction ID", "Is Complete"
        ])

        for order in queryset:
            writer.writerow([
                order.id,
                smart_str(order.user),
                smart_str(order.stall),
                order.pickup_time,
                order.total_cost,
                order.transaction_id,
                order.is_complete
            ])

        return response

    export_orders_as_csv.short_description = "Export selected orders to CSV"

    def export_orders_by_day_excel(self, request, queryset):
        if not request.user.is_superuser:
            queryset = queryset.filter(stall__owner=request.user)

        wb = Workbook()
        wb.remove(wb.active)

        grouped_orders = {}
        for order in queryset:
            date_str = order.pickup_time.strftime('%Y-%m-%d')
            grouped_orders.setdefault(date_str, []).append(order)

        for date_str, orders in grouped_orders.items():
            ws = wb.create_sheet(title=date_str)
            headers = [
                "Order ID", "User", "Stall", "Pickup Time",
                "Total Cost", "Transaction ID", "Is Complete"
            ]
            ws.append(headers)

            for order in orders:
                ws.append([
                    order.id,
                    smart_str(order.user),
                    smart_str(order.stall),
                    order.pickup_time.strftime('%Y-%m-%d %H:%M'),
                    float(order.total_cost),
                    order.transaction_id,
                    "Yes" if order.is_complete else "No"
                ])

            for col in range(1, len(headers)+1):
                ws.column_dimensions[get_column_letter(col)].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"orders_by_day_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    export_orders_by_day_excel.short_description = "Export orders grouped by day (Excel)"

admin_site.register(CustomUser, CustomUserAdmin)
admin_site.register(Stall, StallAdmin)
admin_site.register(MenuItem, MenuItemAdmin)
admin_site.register(Voucher, VoucherAdmin)
admin_site.register(CartItem, CartItemAdmin)
admin_site.register(OrderItem, OrderItemAdmin)
admin_site.register(Order, OrderAdmin)
